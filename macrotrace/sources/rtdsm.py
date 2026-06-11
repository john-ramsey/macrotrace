"""
Philadelphia Fed Real-Time Data Set for Macroeconomists (RTDSM) source.

RTDSM has no API. The Philadelphia Fed publishes, for each macroeconomic
series, a single spreadsheet containing the complete history of every
vintage: rows are observation dates and columns are vintages (the data as
they were known at successive points in time). This source downloads those
spreadsheets and stores every vintage in the macrotrace data model, with no
requirement to keep the Excel files on disk (an optional ``excel_dir`` lets
callers archive them if they wish).

File-naming convention (confirmed across all 115 standard series), where the
leading token is the series ``dataset_id`` (the Philadelphia Fed's mnemonic,
e.g. ``ROUTPUT``):

    {DATASET_ID}{V}v{D}d.xlsx

where ``V`` is the *vintage* frequency (``M`` or ``Q``) and ``D`` is the
*data*/observation frequency (``M`` or ``Q``); the two are independent. The
``series_key`` selects the vintage frequency, e.g. ``{"frequency": "Q"}`` for
quarterly vintages or ``{"frequency": "M"}`` for monthly vintages. The data
frequency is fixed per series and is looked up from the bundled catalog, so we
never probe the server to discover a filename.

Vintage labels map to calendar dates per the Philadelphia Fed documentation:
a quarterly vintage ``YYYY:Qq`` is dated to the middle (15th) of the middle
month of the quarter (February, May, August, November); a monthly vintage
``YYYY:Mm`` is dated to the middle (15th) of month ``m``.

To respect the provider (the files refresh only at the start of the month), the request
cache for a file is set to expire at the start of the next calendar month, so
repeated loads within the same month are served from the local cache without
contacting philadelphiafed.org.
"""

import io
import math
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pytz
import openpyxl
import requests
import requests_cache
from tenacity import (
    before_sleep_log,
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)
from tqdm import tqdm

from macrotrace.models.db import (
    DatasetDimension,
    Observation,
    Release,
    ReleaseDimension,
)
from macrotrace.sources.base import (
    APIClient,
    DatasetManager,
    ObservationManager,
    ReleaseManager,
    SeriesManager,
    UpdateManager,
    UpdateState,
)

import logging

logger = logging.getLogger(__name__)

RTDSM_SOURCE = "RTDSM"
RTDSM_BASE_URL = (
    "https://www.philadelphiafed.org/-/media/FRBP/Assets/"
    "Surveys-And-Data/real-time-data/data-files/xlsx/"
)
UTC = pytz.timezone("UTC")

# Pandas date-offset string for each data (observation) frequency. These must
# be valid offsets because DatasetDimension.frequency is validated against
# pandas (see macrotrace.models.db.FrequencyField).
_PANDAS_FREQ = {"Q": "QS", "M": "MS"}

# Middle month of each quarter (the vintage reference month) and the start
# month of each quarter (the observation reference month).
_QUARTER_MIDDLE_MONTH = {1: 2, 2: 5, 3: 8, 4: 11}
_QUARTER_START_MONTH = {1: 1, 2: 4, 3: 7, 4: 10}


@dataclass
class RTDSMSeries:
    """
    One series' static metadata in the RTDSM catalog.

    Attributes:
        title: Human-readable series title.
        data_freq: Observation frequency, "Q" or "M" (fixed per series).
        vintage_freqs: The vintage frequencies the series is published at, a
            subset of ("Q", "M").
    """

    title: str
    data_freq: str
    vintage_freqs: Tuple[str, ...]


# Catalog of the 115 standard vintage-matrix series. Derived from the
# Philadelphia Fed series pages and verified against the published filenames.
# vintage_freqs is the set of vintage frequencies offered for the series; the
# data_freq is the observation frequency (fixed per series).
RTDSM_CATALOG: Dict[str, RTDSMSeries] = {
    "BASEBASA": RTDSMSeries("Monetary Base (BASEBASA)", "M", ("Q",)),
    "CPI": RTDSMSeries("Consumer Price Index, Quarterly Vintages (CPI)", "M", ("Q",)),
    "CUM": RTDSMSeries("Manufacturing (CUM)", "M", ("M",)),
    "CUT": RTDSMSeries("Total (CUT)", "M", ("M",)),
    "DIV": RTDSMSeries("Dividends (DIV)", "Q", ("Q",)),
    "EMPLOY": RTDSMSeries("Nonfarm Payroll Employment (EMPLOY)", "M", ("M",)),
    "H": RTDSMSeries("Total (H)", "M", ("M",)),
    "HG": RTDSMSeries("Goods-Producing (HG)", "M", ("M",)),
    "HS": RTDSMSeries("Service-Producing (HS)", "M", ("M",)),
    "HSTARTS": RTDSMSeries("Housing Starts (HSTARTS)", "M", ("M",)),
    "IPM": RTDSMSeries("Manufacturing (IPM)", "M", ("M",)),
    "IPT": RTDSMSeries("Total (IPT)", "M", ("M",)),
    "LFC": RTDSMSeries("Civilian Labor Force (LFC)", "M", ("M",)),
    "LFPART": RTDSMSeries("Participation Rate, Constructed (LFPART)", "M", ("M",)),
    "M1": RTDSMSeries("M1 Money Stock (M1)", "M", ("Q",)),
    "M2": RTDSMSeries("M2 Money Stock (M2)", "M", ("Q",)),
    "NBRBASA": RTDSMSeries("Nonborrowed Reserves (NBRBASA)", "M", ("Q",)),
    "NBRECBASA": RTDSMSeries(
        "Nonborrowed Reserves Plus Extended Credit (NBRECBASA)", "M", ("Q",)
    ),
    "NCON": RTDSMSeries(
        "Nominal Personal Consumption Expenditures (NCON)", "Q", ("Q",)
    ),
    "NCONG": RTDSMSeries("Goods (NCONG)", "Q", ("M", "Q")),
    "NCONGM": RTDSMSeries(
        "Personal Consumption Expenditures: Goods (NCONGM)", "M", ("M",)
    ),
    "NCONHH": RTDSMSeries(
        "Household Consumption Expenditures (NCONHH)", "Q", ("M", "Q")
    ),
    "NCONHHM": RTDSMSeries("Household Consumption Expenditures (NCONHHM)", "M", ("M",)),
    "NCONSHH": RTDSMSeries(
        "Household Consumption Expenditures for Services (NCONSHH)", "Q", ("M", "Q")
    ),
    "NCONSHHM": RTDSMSeries(
        "Household Consumption Expdenditures for Services (NCONSHHM)", "M", ("M",)
    ),
    "NCONSNP": RTDSMSeries(
        "Final Consumption Expenditures of NPISH (NCONSNP)", "Q", ("M", "Q")
    ),
    "NCONSNPM": RTDSMSeries(
        "Final Consumption Expenditures of NPISH (NCONSNPM)", "M", ("M",)
    ),
    "NCPROFAT": RTDSMSeries(
        "Nominal Corporate Profits After Tax Without IVA/CCAdj (NCPROFAT)",
        "Q",
        ("M", "Q"),
    ),
    "NCPROFATW": RTDSMSeries(
        "Nominal Corporate Profits After Tax With IVA/CCAdj (NCPROFATW)",
        "Q",
        ("M", "Q"),
    ),
    "NDPI": RTDSMSeries("Nominal Disposable Personal Income (NDPI)", "Q", ("Q",)),
    "NOUTPUT": RTDSMSeries("Nominal GNP/GDP (NOUTPUT)", "Q", ("M", "Q")),
    "NPI": RTDSMSeries("Nominal Personal Income (NPI)", "Q", ("Q",)),
    "NPSAV": RTDSMSeries("Nominal Personal Saving (NPSAV)", "Q", ("Q",)),
    "OLI": RTDSMSeries("Other Labor Income (OLI)", "Q", ("Q",)),
    "OPH": RTDSMSeries("Output Per Hour (OPH)", "Q", ("M", "Q")),
    "P": RTDSMSeries("Price Index for GNP/GDP (P)", "Q", ("M", "Q")),
    "PCON": RTDSMSeries(
        "Price Index for Personal Consumption Expenditures, Constructed (PCON)",
        "Q",
        ("Q",),
    ),
    "PCONG": RTDSMSeries("Goods (PCONG)", "Q", ("M", "Q")),
    "PCONGM": RTDSMSeries(
        "Personal Consumption Expenditures: Goods (PCONGM)", "M", ("M",)
    ),
    "PCONHH": RTDSMSeries(
        "Household Consumption Expenditures (PCONHH)", "Q", ("M", "Q")
    ),
    "PCONHHM": RTDSMSeries("Household Consumption Expenditures (PCONHHM)", "M", ("M",)),
    "PCONSHH": RTDSMSeries(
        "Household Consumption Expenditures for Services (PCONSHH)", "Q", ("M", "Q")
    ),
    "PCONSHHM": RTDSMSeries(
        "Household Consumption Expenditures for Services (PCONSHHM)", "M", ("M",)
    ),
    "PCONSNP": RTDSMSeries(
        "Final Consumption Expenditures of NPISH (PCONSNP)", "Q", ("M", "Q")
    ),
    "PCONSNPM": RTDSMSeries(
        "Final Consumption Expenditures of NPISH (PCONSNPM)", "M", ("M",)
    ),
    "PCONX": RTDSMSeries(
        "Core Price Index for Personal Consumption Expenditures (PCONX)",
        "Q",
        ("M", "Q"),
    ),
    "PCPI": RTDSMSeries("Consumer Price Index, Monthly Vintages (PCPI)", "M", ("M",)),
    "PCPIX": RTDSMSeries("Core Consumer Price Index (PCPIX)", "M", ("M",)),
    "PIMP": RTDSMSeries(
        "Price Index for Imports of Goods and Services (PIMP)", "Q", ("Q",)
    ),
    "PINTI": RTDSMSeries("Personal Interest Income (PINTI)", "Q", ("Q",)),
    "PINTPAID": RTDSMSeries("Interest Paid by Consumers (PINTPAID)", "Q", ("Q",)),
    "POP": RTDSMSeries("Civilian Noninstitutional Population (POP)", "M", ("M",)),
    "PPPI": RTDSMSeries("Producer Price Index, Finished Goods (PPPI)", "M", ("M",)),
    "PPPIX": RTDSMSeries(
        "Core Producer Price Index, Finished Goods (PPPIX)", "M", ("M",)
    ),
    "PROPI": RTDSMSeries("Proprietors' Income (PROPI)", "Q", ("Q",)),
    "PTAX": RTDSMSeries("Personal Tax & Nontax Payments (PTAX)", "Q", ("Q",)),
    "RATESAV": RTDSMSeries("Personal Saving Rate, Constructed (RATESAV)", "Q", ("Q",)),
    "RCON": RTDSMSeries("Total (RCON)", "Q", ("M", "Q")),
    "RCOND": RTDSMSeries("Durable Goods (RCOND)", "Q", ("M", "Q")),
    "RCONDM": RTDSMSeries("Durables (RCONDM)", "M", ("M",)),
    "RCONG": RTDSMSeries("Goods (RCONG)", "Q", ("M", "Q")),
    "RCONGM": RTDSMSeries("Goods (RCONGM)", "M", ("M",)),
    "RCONHH": RTDSMSeries(
        "Household Consumption Expenditures (RCONHH)", "Q", ("M", "Q")
    ),
    "RCONHHM": RTDSMSeries("Household Consumption Expenditures (RCONHHM)", "M", ("M",)),
    "RCONM": RTDSMSeries(
        "Real Personal Consumption Expenditures: Total (RCONM)", "M", ("M",)
    ),
    "RCONND": RTDSMSeries("Nondurable Goods (RCONND)", "Q", ("M", "Q")),
    "RCONNDM": RTDSMSeries("Nondurables (RCONNDM)", "M", ("M",)),
    "RCONS": RTDSMSeries("Services (RCONS)", "Q", ("M", "Q")),
    "RCONSHH": RTDSMSeries(
        "Real Household Consumption Expenditures for Services (RCONSHH)",
        "Q",
        ("M", "Q"),
    ),
    "RCONSHHM": RTDSMSeries(
        "Household Consumption Expenditures for Services (RCONSHHM)", "M", ("M",)
    ),
    "RCONSM": RTDSMSeries("Services (RCONSM)", "M", ("M",)),
    "RCONSNP": RTDSMSeries(
        "Real Final Consumption Expenditures of NPISH (RCONSNP)", "Q", ("M", "Q")
    ),
    "RCONSNPM": RTDSMSeries(
        "Final Consumption Expenditures of NPISH (RCONSNPM)", "M", ("M",)
    ),
    "RENTI": RTDSMSeries("Rental Income of Persons (RENTI)", "Q", ("Q",)),
    "REX": RTDSMSeries("Real Exports of Goods and Services (REX)", "Q", ("M", "Q")),
    "RG": RTDSMSeries("Total (RG)", "Q", ("M", "Q")),
    "RGF": RTDSMSeries("Federal (RGF)", "Q", ("M", "Q")),
    "RGSL": RTDSMSeries("State and Local (RGSL)", "Q", ("M", "Q")),
    "RIMP": RTDSMSeries("Real Imports of Goods and Services (RIMP)", "Q", ("M", "Q")),
    "RINVBF": RTDSMSeries("Nonresidential (RINVBF)", "Q", ("M", "Q")),
    "RINVCHI": RTDSMSeries("Change in Private Inventories (RINVCHI)", "Q", ("M", "Q")),
    "RINVRESID": RTDSMSeries("Residential (RINVRESID)", "Q", ("M", "Q")),
    "RNX": RTDSMSeries("Real Net Exports of Goods and Services (RNX)", "Q", ("M", "Q")),
    "ROUTPUT": RTDSMSeries("Real GNP/GDP (ROUTPUT)", "Q", ("M", "Q")),
    "RUC": RTDSMSeries("Unemployment Rate (RUC)", "M", ("Q",)),
    "SSCONTRIB": RTDSMSeries(
        "Personal Contributions for Social Insurance (SSCONTRIB)", "Q", ("Q",)
    ),
    "TRANPF": RTDSMSeries(
        "Personal Transfer Payments to Foreigners (TRANPF)", "Q", ("Q",)
    ),
    "TRANR": RTDSMSeries("Transfer Payments (TRANR)", "Q", ("Q",)),
    "TRBASA": RTDSMSeries("Total Reserves (TRBASA)", "M", ("Q",)),
    "ULC": RTDSMSeries("Unit Labor Costs (ULC)", "Q", ("M", "Q")),
    "WSD": RTDSMSeries("Wage and Salary Disbursements (WSD)", "Q", ("Q",)),
    "YNCFC": RTDSMSeries("Consumption of Fixed Capital (YNCFC)", "Q", ("M", "Q")),
    "YNCFCG": RTDSMSeries(
        "Consumption of Fixed Capital, Government (YNCFCG)", "Q", ("M", "Q")
    ),
    "YNCFCP": RTDSMSeries(
        "Consumption of Fixed Capital, Private (YNCFCP)", "Q", ("M", "Q")
    ),
    "YNCOMPEP": RTDSMSeries(
        "Compensation of Employees, Paid (YNCOMPEP)", "Q", ("M", "Q")
    ),
    "YNCPRFATW": RTDSMSeries(
        "Corporate Profits After Tax with IVA and CCA (YNCPRFATW)", "Q", ("M", "Q")
    ),
    "YNCPRFW": RTDSMSeries(
        "Corporate Profits with IVA and CCA (YNCPRFW)", "Q", ("M", "Q")
    ),
    "YNCTAX": RTDSMSeries("Taxes on Corporate Income (YNCTAX)", "Q", ("M", "Q")),
    "YNDPAID": RTDSMSeries("Net Dividends Paid (YNDPAID)", "Q", ("M", "Q")),
    "YNGDI": RTDSMSeries("Nominal GDI (YNGDI)", "Q", ("M", "Q")),
    "YNGSUB": RTDSMSeries("Government Subsidies (YNGSUB)", "Q", ("M", "Q")),
    "YNIPAID": RTDSMSeries(
        "Net Interest and Miscellaneous Payments (YNIPAID)", "Q", ("M", "Q")
    ),
    "YNOS": RTDSMSeries("Net Operating Surplus (YNOS)", "Q", ("M", "Q")),
    "YNOSG": RTDSMSeries("Net Operating Surplus, Government (YNOSG)", "Q", ("M", "Q")),
    "YNOSP": RTDSMSeries("Net Operating Surplus, Private (YNOSP)", "Q", ("M", "Q")),
    "YNPINCW": RTDSMSeries(
        "Proprietors' Income with IVA and CCA (YNPINCW)", "Q", ("M", "Q")
    ),
    "YNRINC": RTDSMSeries("Rental Income with CCA (YNRINC)", "Q", ("M", "Q")),
    "YNSD": RTDSMSeries("Statistical Discrepancy (YNSD)", "Q", ("M", "Q")),
    "YNSWS": RTDSMSeries("Supplements to Wages and Salaries (YNSWS)", "Q", ("M", "Q")),
    "YNTAXR": RTDSMSeries("Taxes on Production and Imports (YNTAXR)", "Q", ("M", "Q")),
    "YNTRPAY": RTDSMSeries(
        "Business Current Transfer Payments, Net (YNTRPAY)", "Q", ("M", "Q")
    ),
    "YNUCPRFW": RTDSMSeries(
        "Undistributed Corporate Profits with IVA and CCA (YNUCPRFW)", "Q", ("M", "Q")
    ),
    "YNWS": RTDSMSeries("Wages and Salaries (YNWS)", "Q", ("M", "Q")),
    "YPDGDP": RTDSMSeries("Implicit Price Deflator, GDP (YPDGDP)", "Q", ("M", "Q")),
    "YRGDI": RTDSMSeries(
        "Real GDI [computed as Nominal GDI divided by Nominal GDP/Real GDP] (YRGDI)",
        "Q",
        ("M", "Q"),
    ),
}


@dataclass
class ParsedVintageFile:
    """
    Structured contents of one parsed vintage spreadsheet.

    Attributes:
        vintages: (vintage_label, release_date) pairs in column order.
        cells: Maps each vintage label to its non-missing
            (observation_timestamp, value) pairs (#N/A cells are omitted).
    """

    vintages: List[Tuple[str, datetime]]
    cells: Dict[str, List[Tuple[datetime, float]]]


def _first_of_next_month(now: datetime) -> datetime:
    """
    Return midnight UTC on the first day of the month after ``now``.

    Used as the request-cache expiry so a file fetched at any point in a month
    is served from cache for the rest of that month and refreshed once the next
    month begins (RTDSM files only update at month-end).
    """
    year = now.year + 1 if now.month == 12 else now.year
    month = 1 if now.month == 12 else now.month + 1
    return datetime(year, month, 1, tzinfo=UTC)


def _resolve_vintage_freq(
    dataset_id: str, info: RTDSMSeries, requested: Optional[str]
) -> str:
    """
    Resolve the vintage frequency to use for a series.

    If ``requested`` is None, default to the series' only vintage frequency, or
    to quarterly when both are offered. Otherwise validate the request against
    what the series actually offers.

    Args:
        dataset_id: The uppercase series identifier.
        info: The catalog entry for the series.
        requested: The frequency requested via the series key, or None.

    Returns:
        str: The resolved vintage frequency, "Q" or "M".

    Raises:
        ValueError: If ``requested`` is not "Q"/"M" or is not offered.
    """
    available = info.vintage_freqs
    if requested is None:
        resolved = "Q" if "Q" in available else available[0]
        logger.debug(
            f"No frequency requested for {dataset_id}; defaulting to {resolved} "
            f"(available: {', '.join(available)})"
        )
        return resolved

    req = str(requested).strip().upper()
    if req not in ("M", "Q"):
        raise ValueError(f"RTDSM frequency must be 'Q' or 'M', got {requested!r}.")
    if req not in available:
        raise ValueError(
            f"RTDSM series {dataset_id} does not offer {req}-frequency vintages. "
            f"Available: {', '.join(available)}."
        )
    return req


def _build_filename(dataset_id: str, vintage_freq: str, data_freq: str) -> str:
    """
    Build the RTDSM spreadsheet filename for a series and vintage frequency.

    The Philadelphia Fed media server is case-insensitive on the filename, so
    we emit a fully lowercase name regardless of the (inconsistent) casing used
    in the site's own links.

    Args:
        dataset_id: The series identifier (e.g. ``ROUTPUT``).
        vintage_freq: The vintage frequency, "Q" or "M".
        data_freq: The data (observation) frequency, "Q" or "M".

    Returns:
        str: The lowercase filename, e.g. "routputqvqd.xlsx".
    """
    return f"{dataset_id}{vintage_freq}v{data_freq}d.xlsx".lower()


def _parse_vintage_label(
    label: str, dataset_id: str, current_year: Optional[int] = None
) -> datetime:
    """
    Parse a vintage column header into its release date.

    Headers look like ``ROUTPUT65Q4`` (quarterly vintage) or ``RCON65M11``
    (monthly vintage): the dataset_id, a two-digit year, ``Q``/``M``, and a
    period number. Quarterly vintages are dated to the 15th of the middle month
    of the quarter; monthly vintages to the 15th of the month.

    The two-digit year is resolved relative to the current year rather than a
    fixed pivot: real-time vintages can never be dated in the future, so the
    1900s reading is chosen whenever the 2000s reading would be. This correctly
    dates monthly-vintage series that begin before 1965 (e.g. IPM/IPT start at
    1962:M11 and EMPLOY at 1964:M12) while still dating recent vintages in the
    2000s. It is unambiguous because no RTDSM series spans 100 years.

    Args:
        label: The vintage column header.
        dataset_id: The series identifier, used to strip the label's prefix.
        current_year: The reference year for two-digit-year resolution;
            defaults to the current UTC year.

    Returns:
        datetime: The timezone-aware (UTC) release date.

    Raises:
        ValueError: If the label does not match the expected pattern.
    """
    if current_year is None:
        current_year = datetime.now(UTC).year

    text = str(label).strip().upper()
    prefix = dataset_id.upper()
    suffix = text[len(prefix) :] if text.startswith(prefix) else ""

    # After the dataset_id the label is <yy><M|Q><period>: a two-digit year, a
    # frequency letter, then a one or two digit period number (e.g. the "65Q4"
    # in ROUTPUT65Q4, or the "65M11" in RCON65M11).
    year_str, freq_letter, number_str = suffix[:2], suffix[2:3], suffix[3:]
    if (
        not year_str.isdigit()
        or freq_letter not in ("M", "Q")
        or not number_str.isdigit()
        or len(number_str) > 2
    ):
        raise ValueError(
            f"Unrecognized RTDSM vintage label {label!r} for series {dataset_id}."
        )
    yy = int(year_str)
    period = freq_letter
    number = int(number_str)
    # Prefer the 2000s reading, but fall back to the 1900s whenever that would
    # place the vintage in the future (a small +1 margin allows a vintage
    # labeled with next year near a year boundary).
    year = 2000 + yy
    if year > current_year + 1:
        year -= 100

    if period == "Q":
        if number not in _QUARTER_MIDDLE_MONTH:
            raise ValueError(f"Invalid quarter in vintage label {label!r}.")
        month = _QUARTER_MIDDLE_MONTH[number]
    else:
        if not 1 <= number <= 12:
            raise ValueError(f"Invalid month in vintage label {label!r}.")
        month = number

    return datetime(year, month, 15, tzinfo=UTC)


def _parse_observation_label(label: str, data_freq: str) -> datetime:
    """
    Parse an observation row label into its timestamp.

    Quarterly rows are labeled ``YYYY:Qq`` and anchored to the first day of the
    quarter; monthly rows are labeled ``YYYY:MM`` and anchored to the first day
    of the month.

    Args:
        label: The observation row label.
        data_freq: The data frequency, "Q" or "M".

    Returns:
        datetime: The timezone-aware (UTC) observation timestamp.

    Raises:
        ValueError: If the label does not match the expected pattern.
    """
    # Labels are "<year>:<period>"; split on the single colon.
    year_str, sep, period_str = str(label).strip().partition(":")
    valid_year = sep == ":" and len(year_str) == 4 and year_str.isdigit()

    if data_freq == "Q":
        # Quarterly period is "Q" (any case) followed by a quarter digit 1-4.
        if not (
            valid_year
            and len(period_str) == 2
            and period_str[0].upper() == "Q"
            and period_str[1] in "1234"
        ):
            raise ValueError(
                f"Unrecognized RTDSM quarterly observation label {label!r}."
            )
        month = _QUARTER_START_MONTH[int(period_str[1])]
        return datetime(int(year_str), month, 1, tzinfo=UTC)

    # Monthly period is a one or two digit month number.
    if not (valid_year and period_str.isdigit() and len(period_str) <= 2):
        raise ValueError(f"Unrecognized RTDSM monthly observation label {label!r}.")
    month = int(period_str)
    if not 1 <= month <= 12:
        raise ValueError(f"Invalid month in observation label {label!r}.")
    return datetime(int(year_str), month, 1, tzinfo=UTC)


def _coerce_value(value: Any) -> Optional[float]:
    """
    Coerce a spreadsheet cell value to a float, or None if missing.

    RTDSM marks unavailable cells with the literal text ``#N/A``; any
    non-numeric cell (or NaN) is treated as missing and dropped.

    Args:
        value: The raw cell value from openpyxl.

    Returns:
        Optional[float]: The numeric value, or None if the cell is missing.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number):
            return None
        return number
    return None


def _parse_workbook(
    content: bytes, dataset_id: str, data_freq: str
) -> ParsedVintageFile:
    """
    Parse a downloaded RTDSM spreadsheet into vintages and observations.

    Args:
        content: The raw .xlsx bytes.
        dataset_id: The series identifier (e.g. ``ROUTPUT``).
        data_freq: The data frequency, "Q" or "M".

    Returns:
        ParsedVintageFile: Parsed vintages and their non-missing observations.

    Raises:
        ValueError: If the workbook has no header row.
    """
    workbook = openpyxl.load_workbook(
        io.BytesIO(content), read_only=True, data_only=True
    )
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError("RTDSM workbook contains no rows.")

        # Column A is "DATE"; columns B+ are vintage identifiers. Keep the
        # column index alongside the label so we are robust to any gaps.
        vintage_columns = [
            (index, str(label))
            for index, label in enumerate(header)
            if index >= 1 and label not in (None, "")
        ]

        current_year = datetime.now(UTC).year
        vintages = []
        cells: Dict[str, List] = {}
        for _index, label in vintage_columns:
            vintages.append(
                (label, _parse_vintage_label(label, dataset_id, current_year))
            )
            cells[label] = []

        for row in rows:
            if not row:
                continue
            obs_label = row[0]
            if obs_label in (None, ""):
                continue
            obs_timestamp = _parse_observation_label(str(obs_label), data_freq)
            for index, label in vintage_columns:
                value = row[index] if index < len(row) else None
                number = _coerce_value(value)
                if number is None:
                    continue
                cells[label].append((obs_timestamp, number))
    finally:
        workbook.close()

    return ParsedVintageFile(vintages=vintages, cells=cells)


def _ensure_utc(dt: Optional[datetime]) -> Optional[datetime]:
    """
    Return the datetime as a timezone-aware UTC value, or None.

    RTDSM release dates are stored in UTC, so any caller-supplied window bound
    must also be timezone-aware to compare correctly.
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


class RTDSMAPIClient(APIClient):
    """
    Downloads and parses a single RTDSM spreadsheet.

    One client is created per (series, vintage frequency). The parsed workbook
    is memoized so the managers that make up one update share a single download
    and parse.
    """

    def __init__(
        self,
        dataset_id: str,
        filename: str,
        data_freq: str,
        excel_dir: Optional[str] = None,
        cache_settings: Optional[Dict[str, Any]] = None,
        cache_path: Optional[str] = None,
    ):
        self.dataset_id = dataset_id
        self.filename = filename
        self.data_freq = data_freq
        self.excel_dir = excel_dir
        self._parsed: Optional[ParsedVintageFile] = None
        super().__init__(
            base_url=RTDSM_BASE_URL,
            cache_settings=cache_settings,
            cache_path=cache_path,
        )

    def _get_request_headers(self) -> Dict[str, Any]:
        """RTDSM downloads need no special headers beyond the user agent."""
        return {}

    def _get_default_params(self) -> Dict[str, str]:
        """RTDSM downloads need no query parameters."""
        return {}

    @retry(
        stop=stop_after_attempt(4),
        wait=wait_exponential(max=30),
        before_sleep=before_sleep_log(logger, logging.WARNING),
        retry=retry_if_exception_type(requests.RequestException),
        reraise=True,
    )
    def _download(self) -> bytes:
        """
        Download the spreadsheet bytes, honoring the month-aligned cache.

        Returns:
            bytes: The raw .xlsx content.

        Raises:
            ValueError: If the response is not a valid .xlsx file (the media
                server returns an HTML error page with HTTP 200 for missing
                files, so content is validated by its zip magic bytes).
        """
        url = self.base_url + self.filename
        headers = {"User-Agent": self.user_agent}
        logger.info(f"Downloading RTDSM file: {self.filename}")

        if isinstance(self.session, requests_cache.CachedSession):
            # Expire the cached copy at the start of next month so repeated
            # loads within a month never re-request the provider's server.
            expire_at = _first_of_next_month(datetime.now(UTC))
            response = self.session.get(url, headers=headers, expire_after=expire_at)
        else:
            response = self.session.get(url, headers=headers)

        is_cached = getattr(response, "from_cache", False)
        response.raise_for_status()
        content = response.content
        logger.debug(
            f"RTDSM download {self.filename}: status={response.status_code}, "
            f"cached={is_cached}, size={len(content)} bytes"
        )

        if content[:4] != b"PK\x03\x04":
            raise ValueError(
                f"RTDSM download for {self.filename} did not return a valid "
                f".xlsx file ({len(content)} bytes). The series may not exist "
                f"at the requested vintage frequency, or the provider returned "
                f"an error page."
            )

        if self.excel_dir:
            self._save_excel(content)

        return content

    def _save_excel(self, content: bytes) -> None:
        """
        Write the downloaded spreadsheet to ``excel_dir`` if requested.

        Args:
            content: The raw .xlsx content.
        """
        os.makedirs(self.excel_dir, exist_ok=True)
        path = os.path.join(self.excel_dir, self.filename)
        with open(path, "wb") as handle:
            handle.write(content)
        logger.info(f"Saved RTDSM spreadsheet to {path}")

    def get_parsed_file(self) -> ParsedVintageFile:
        """
        Return the parsed workbook, downloading and parsing once.

        Returns:
            ParsedVintageFile: The parsed vintages and observations.

        Raises:
            ValueError: If the download does not return a valid .xlsx file.
        """
        if self._parsed is None:
            content = self._download()
            self._parsed = _parse_workbook(content, self.dataset_id, self.data_freq)
            logger.info(
                f"Parsed RTDSM file {self.filename}: "
                f"{len(self._parsed.vintages)} vintage(s)"
            )
        return self._parsed


class RTDSMDatasetManager(DatasetManager):
    def __init__(self, api_client: RTDSMAPIClient):
        super().__init__(api_client)

    def fetch_new_dataset_dimensions(
        self, state: UpdateState
    ) -> List[DatasetDimension]:
        """
        Create the single numeric dimension that defines the series.

        Unlike FRED (which versions its one dimension by realtime period) an
        RTDSM series has a single, static definition: one numeric dimension
        spanning every vintage. We create it once, on first load, with a
        ``valid_from`` early enough to cover every vintage in the file so all
        releases associate with it.

        Args:
            state: The update state.

        Returns:
            List[DatasetDimension]: The new dimension, or an empty list if it
                already exists or there are no releases to anchor it.
        """
        existing = self._get_all_local_dataset_dimensions(state.dataset.id)
        if existing:
            logger.debug(
                f"RTDSM dimension already exists for {self.api_client.dataset_id}; "
                f"no new dimensions."
            )
            return []

        parsed = self.api_client.get_parsed_file()
        if not parsed.vintages:
            logger.debug("No vintages found in RTDSM file; no dimension created.")
            return []

        earliest_release = min(release_date for _, release_date in parsed.vintages)
        info = RTDSM_CATALOG[self.api_client.dataset_id]
        dimension = DatasetDimension(
            dataset=state.dataset,
            dataset_dimension_id=self.api_client.dataset_id,
            title=info.title,
            type="numeric",
            frequency=_PANDAS_FREQ[self.api_client.data_freq],
            description=None,
            units=None,
            seasonal_adjustment=None,
            valid_from=earliest_release,
            valid_to=None,
        )
        logger.info(
            f"Created RTDSM dataset dimension for {self.api_client.dataset_id} "
            f"(valid_from={earliest_release})"
        )
        return [dimension]


class RTDSMReleaseManager(ReleaseManager):
    def __init__(self, api_client: RTDSMAPIClient):
        super().__init__(api_client)

    def fetch_new_releases(self, state: UpdateState) -> List[Release]:
        """
        Create a Release for each vintage column not already stored.

        The whole vintage history is downloaded, so releases are filtered
        client-side against any requested vintage window and against what is
        already in the database.

        Args:
            state: The update state.

        Returns:
            List[Release]: The new releases.
        """
        state.release_start_date = _ensure_utc(state.release_start_date)
        state.release_end_date = _ensure_utc(state.release_end_date)

        parsed = self.api_client.get_parsed_file()
        current_release_dates = self._get_current_releases_in_db(state.dataset.id)

        new_releases = []
        for _label, release_date in parsed.vintages:
            if self._is_new_release(
                release_date, current_release_dates
            ) and self._is_wanted_release(
                release_date, state.release_start_date, state.release_end_date
            ):
                new_releases.append(
                    Release(dataset=state.dataset, release_date=release_date)
                )

        logger.info(
            f"Found {len(new_releases)} new RTDSM release(s) out of "
            f"{len(parsed.vintages)} vintage(s)"
        )
        return new_releases

    def fetch_new_release_dimensions(
        self, state: UpdateState
    ) -> List[ReleaseDimension]:
        """
        Associate each new release with the series' single dimension.

        Args:
            state: The update state.

        Returns:
            List[ReleaseDimension]: The new release-dimension associations.

        Raises:
            ValueError: If the dataset has no dimension to associate.
        """
        all_dims = self._get_all_local_dataset_dimensions(state.dataset.id)
        if not all_dims:
            raise ValueError(
                f"Dataset {state.dataset.id} has no dimensions to associate "
                f"with releases."
            )

        new_release_dimensions = []
        for release in state.new_releases:
            for dimension in all_dims:
                in_lower_bound = release.release_date >= dimension.valid_from
                in_upper_bound = (
                    dimension.valid_to is None
                    or release.release_date <= dimension.valid_to
                )
                if in_lower_bound and in_upper_bound:
                    new_release_dimensions.append(
                        ReleaseDimension(release=release, dimension=dimension)
                    )

        logger.info(
            f"Created {len(new_release_dimensions)} RTDSM release-dimension "
            f"association(s)"
        )
        return new_release_dimensions


class RTDSMSeriesManager(SeriesManager):
    def __init__(self, api_client: RTDSMAPIClient):
        super().__init__(api_client)

    def fetch_new_series_dimension_filters(self, state: UpdateState) -> List:
        """
        RTDSM series have no dimension filters.

        The ``frequency`` entry in the series key selects which spreadsheet
        (vintage cadence) to download; it is not a dataset dimension, so there
        are no SeriesDimensionFilter rows to create. The base implementation
        would try to look up a dimension named "frequency" and fail, so we
        override it to return nothing.

        Args:
            state: The update state.

        Returns:
            List: Always empty.
        """
        return []


class RTDSMObservationManager(ObservationManager):
    def __init__(self, api_client: RTDSMAPIClient):
        super().__init__(api_client)

    def fetch_new_observations(self, state: UpdateState) -> List[Observation]:
        """
        Create observations for every non-missing cell of the new releases.

        Args:
            state: The update state.

        Returns:
            List[Observation]: The new observations.
        """
        if not state.new_releases:
            logger.debug("No new RTDSM releases; no observations to create.")
            return []

        parsed = self.api_client.get_parsed_file()
        date_to_label = {release_date: label for label, release_date in parsed.vintages}

        new_observations = []
        for release in tqdm(
            state.new_releases, desc="Processing RTDSM vintages", leave=False
        ):
            label = date_to_label.get(release.release_date)
            if label is None:
                logger.warning(
                    f"No vintage column found for release "
                    f"{release.release_date}; skipping."
                )
                continue
            for obs_timestamp, value in parsed.cells.get(label, []):
                new_observations.append(
                    Observation(
                        series=state.series,
                        release=release,
                        observation_timestamp=obs_timestamp,
                        value=value,
                    )
                )

        logger.info(f"Created {len(new_observations)} new RTDSM observation(s)")
        return new_observations


class RTDSMUpdateManager(UpdateManager):
    NATIVE_OBSERVATION_TZ = UTC

    def __init__(
        self,
        dataset_id: str,
        source: str = RTDSM_SOURCE,
        series_key: Optional[Dict] = None,
        release_start_date: Optional[datetime] = None,
        release_end_date: Optional[datetime] = None,
        db_path: Optional[str] = None,
        cache_settings: Optional[Dict[str, Any]] = None,
        cache_path: Optional[str] = None,
        excel_dir: Optional[str] = None,
    ):
        """
        Initialize an RTDSM update manager for a single series.

        Args:
            dataset_id: The series identifier (e.g. "ROUTPUT"); case-insensitive.
            source: The source name; defaults to "RTDSM".
            series_key: Optionally ``{"frequency": "Q"}`` or ``{"frequency":
                "M"}`` to select the vintage frequency. If omitted, defaults to
                the series' only vintage frequency, or quarterly when both are
                offered.
            release_start_date: Optional lower bound on vintage dates to load.
            release_end_date: Optional upper bound on vintage dates to load.
            db_path: Optional path to the SQLite database.
            cache_settings: Optional request-cache settings.
            cache_path: Optional path to the request-cache SQLite file.
            excel_dir: Optional directory in which to save the downloaded
                spreadsheet. If None, the file is parsed in memory and not kept.

        Raises:
            ValueError: If ``dataset_id`` is not a known RTDSM series, or if the
                requested vintage frequency is not offered by that series. The
                catalog is bundled, so an unknown series fails fast here rather
                than at fetch time.
        """
        dataset_id = dataset_id.upper()
        self.dataset_id = dataset_id
        self.excel_dir = excel_dir

        info = RTDSM_CATALOG.get(dataset_id)
        if info is None:
            raise ValueError(
                f"Unknown RTDSM series {dataset_id!r}. See "
                f"macrotrace.sources.rtdsm.RTDSM_CATALOG for the "
                f"{len(RTDSM_CATALOG)} supported series identifiers."
            )

        requested = series_key.get("frequency") if series_key else None
        self.vintage_freq = _resolve_vintage_freq(dataset_id, info, requested)
        self.data_freq = info.data_freq
        self.filename = _build_filename(dataset_id, self.vintage_freq, info.data_freq)
        resolved_series_key = {"frequency": self.vintage_freq}
        logger.debug(
            f"Initializing RTDSMUpdateManager for {dataset_id} "
            f"(vintage_freq={self.vintage_freq}, file={self.filename})"
        )

        super().__init__(
            dataset_id=dataset_id,
            source=source,
            series_key=resolved_series_key,
            release_start_date=release_start_date,
            release_end_date=release_end_date,
            db_path=db_path,
            cache_settings=cache_settings,
            cache_path=cache_path,
        )

    def _create_api_client(
        self,
        cache_settings: Optional[Dict[str, Any]] = None,
        cache_path: Optional[str] = None,
    ) -> RTDSMAPIClient:
        return RTDSMAPIClient(
            dataset_id=self.dataset_id,
            filename=self.filename,
            data_freq=self.data_freq,
            excel_dir=self.excel_dir,
            cache_settings=cache_settings,
            cache_path=cache_path,
        )

    def _create_dataset_manager(self) -> DatasetManager:
        return RTDSMDatasetManager(self.api_client)

    def _create_release_manager(self) -> ReleaseManager:
        return RTDSMReleaseManager(self.api_client)

    def _create_series_manager(self) -> SeriesManager:
        return RTDSMSeriesManager(self.api_client)

    def _create_observation_manager(self) -> ObservationManager:
        return RTDSMObservationManager(self.api_client)
